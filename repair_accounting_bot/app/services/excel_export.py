from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.accounting import format_toman

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(bold=True, size=14)
RTL = Alignment(horizontal='right', vertical='center', wrap_text=True)


def _autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        length = max(len(str(cell.value or '')) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length + 2, 12), 40)


def _header_row(sheet, row: int, values: list[str]) -> None:
    for col, value in enumerate(values, start=1):
        cell = sheet.cell(row=row, column=col, value=value)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = RTL


def build_accounting_workbook(
    dashboard: dict[str, Any],
    repairs: list[dict[str, Any]],
    customer_debts: list[dict[str, Any]],
) -> Workbook:
    wb = Workbook()
    summary = wb.active
    summary.title = 'خلاصه'
    summary.sheet_view.rightToLeft = True
    summary['A1'] = 'گزارش حسابداری CTTEL'
    summary['A1'].font = TITLE_FONT
    summary['A2'] = datetime.now().strftime('%Y-%m-%d %H:%M')

    rows = [
        ('پرونده باز', dashboard['open_count']),
        ('سود فروشگاه', dashboard['shop_profit']),
        ('جمع سهم تعمیرکار', dashboard['technician_share']),
        ('پرداخت‌شده به تعمیرکار', dashboard.get('technician_paid', 0)),
        ('مانده طلب تعمیرکار', dashboard.get('technician_debt', 0)),
        ('جمع بدهی مشتری', dashboard['customer_debt']),
        ('جمع بدهی قطعه‌فروش', dashboard['supplier_debt']),
    ]
    start = 4
    _header_row(summary, start, ['شرح', 'مبلغ (تومان)'])
    for idx, (label, amount) in enumerate(rows, start=start + 1):
        summary.cell(row=idx, column=1, value=label).alignment = RTL
        summary.cell(row=idx, column=2, value=amount)

    tech_sheet = wb.create_sheet('تعمیرکاران')
    tech_sheet.sheet_view.rightToLeft = True
    _header_row(tech_sheet, 1, ['نام', 'درصد', 'سهم', 'پرداخت‌شده', 'مانده طلب'])
    for idx, row in enumerate(dashboard.get('technicians', []), start=2):
        tech_sheet.cell(row=idx, column=1, value=row['name']).alignment = RTL
        tech_sheet.cell(row=idx, column=2, value=row['pct'])
        tech_sheet.cell(row=idx, column=3, value=int(row['share']))
        tech_sheet.cell(row=idx, column=4, value=int(row.get('paid') or 0))
        tech_sheet.cell(row=idx, column=5, value=int(row.get('debt') or 0))

    sup_sheet = wb.create_sheet('قطعه‌فروش')
    sup_sheet.sheet_view.rightToLeft = True
    _header_row(sup_sheet, 1, ['فروشنده', 'بدهی (تومان)'])
    for idx, row in enumerate(dashboard.get('suppliers', []), start=2):
        sup_sheet.cell(row=idx, column=1, value=row['name']).alignment = RTL
        sup_sheet.cell(row=idx, column=2, value=int(row['debt']))

    cust_sheet = wb.create_sheet('بدهی مشتریان')
    cust_sheet.sheet_view.rightToLeft = True
    _header_row(cust_sheet, 1, ['مشتری', 'تماس', 'بدهی (تومان)'])
    for idx, row in enumerate(customer_debts, start=2):
        cust_sheet.cell(row=idx, column=1, value=row['name']).alignment = RTL
        cust_sheet.cell(row=idx, column=2, value=row.get('phone') or '—').alignment = RTL
        cust_sheet.cell(row=idx, column=3, value=int(row['debt']))

    repairs_sheet = wb.create_sheet('پرونده‌ها')
    repairs_sheet.sheet_view.rightToLeft = True
    _header_row(
        repairs_sheet,
        1,
        [
            'شماره',
            'مشتری',
            'دستگاه',
            'تعمیرکار',
            'درصد',
            'اجرت',
            'فروش قطعه',
            'جمع',
            'دریافت',
            'بدهی مشتری',
            'بدهی قطعه‌فروش',
            'سهم تعمیرکار',
            'پرداخت تعمیرکار',
            'مانده طلب تعمیرکار',
            'سود مغازه',
        ],
    )
    for idx, repair in enumerate(repairs, start=2):
        totals = repair['totals']
        values = [
            repair['id'],
            repair['customer_name'],
            repair['device'],
            repair.get('technician_name') or '—',
            repair['technician_pct'],
            totals.labor_amount,
            totals.parts_sell,
            totals.customer_total,
            totals.customer_paid,
            totals.customer_debt,
            totals.supplier_debt,
            totals.technician_share,
            totals.technician_paid,
            totals.technician_debt,
            totals.shop_profit,
        ]
        for col, value in enumerate(values, start=1):
            repairs_sheet.cell(row=idx, column=col, value=value).alignment = RTL

    for sheet in wb.worksheets:
        _autosize_columns(sheet)
    return wb


def build_invoice_workbook(repair: dict[str, Any]) -> Workbook:
    wb = Workbook()
    sheet = wb.active
    sheet.title = f"فاکتور {repair['id']}"
    sheet.sheet_view.rightToLeft = True
    totals = repair['totals']

    sheet['A1'] = f"فاکتور فروش #{repair['id']}"
    sheet['A1'].font = TITLE_FONT
    sheet['A2'] = 'CTTEL · سی تی تل'
    sheet['A3'] = datetime.now().strftime('%Y-%m-%d %H:%M')

    info = [
        ('مشتری', repair['customer_name']),
        ('تماس', repair['customer_phone'] or '—'),
        ('دستگاه', repair['device']),
        ('ایراد', repair['issue']),
        ('تعمیرکار', repair.get('technician_name') or '—'),
        ('درصد تعمیرکار', f"{repair['technician_pct']}%"),
    ]
    row = 5
    for label, value in info:
        sheet.cell(row=row, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row, column=2, value=value).alignment = RTL
        row += 1

    row += 1
    _header_row(sheet, row, ['شرح', 'مبلغ (تومان)'])
    row += 1
    sheet.cell(row=row, column=1, value='اجرت تعمیر').alignment = RTL
    sheet.cell(row=row, column=2, value=totals.labor_amount)
    row += 1
    for part in repair['parts']:
        sheet.cell(row=row, column=1, value=part['part_name']).alignment = RTL
        sheet.cell(row=row, column=2, value=int(part['sell_price']))
        row += 1

    row += 1
    for label, amount in [
        ('جمع کل', totals.customer_total),
        ('پرداخت‌شده', totals.customer_paid),
        ('مانده', totals.customer_debt),
    ]:
        sheet.cell(row=row, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row, column=2, value=amount)
        row += 1

    _autosize_columns(sheet)
    return wb


def save_workbook(wb: Workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
